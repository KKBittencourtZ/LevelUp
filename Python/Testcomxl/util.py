# Aqui estão algumas ferramentas que podem se úteis na hora de relacionar as tabelas


def tabela(ws, linha, coluna):
    """

    ws: a worksheet que será printada
    linha: quantas linhas serão printadas
    coluna: quantas colunas serão printadas
    O resultado sairá em uma linha com todos os dados das células
    By<César Bittencourt>

    """
    from openpyxl.utils import get_column_letter
    for lin in range(1, linha + 1):
        for col in range(1, coluna + 1):
            char = get_column_letter(col)
            if col < coluna:
                print(ws[char + str(lin)].value, end=' ')
            elif col == coluna:
                print(ws[char + str(lin)].value)




def salvar_bbtk(wb):
    '''

    wb: a variável que recebeu a função "load_workbook"


    '''
    wb.save("C:\ProjectGitHub\LevelUp\Python\Testcomxl/bbtk.xlsx")



def alterar_em_todas_sheets(ws, x, txt, wb):
    tsheets = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    for c in range(0, len(tsheets)):
        print(c, end=' ')
        ws = wb[tsheets[c]]
        ws[x] = txt


def save_ass(ws, c, msg, wb):
    ws[c] = msg
    wb.save("C:\ProjectGitHub\LevelUp\Python\Testcomxl/bbtk.xlsx")