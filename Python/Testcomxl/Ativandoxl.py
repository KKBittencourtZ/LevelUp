from openpyxl import Workbook, load_workbook
import util

wb = load_workbook("C:\ProjectGitHub\LevelUp\Python\Testcomxl/bbtk.xlsx")



ws = wb['A']



util.tabela(ws, 2, 4)





# ws.title = "A"      # nomear a sheet
# wb.save("C:\ProjectGitHub\LevelUp\Python\Testcomxl/bbtk.xlsx")
# print(ws['A1'].value)
